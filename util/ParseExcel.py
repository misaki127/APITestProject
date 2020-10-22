
import openpyxl
from openpyxl.styles import Border,Side,Font
import time



class ParseExcel(object):

    def __init__(self):
        self.workbook=None
        self.excelFile=None
        self.font=Font(color=None) #设置字体的颜色
        #颜色对应的RGB值
        self.RGBDict={'red':'FFFF3030','green':'FF008B00'}

    #将excel文件加载到内存，并获取其workbook对象
    def loadWorkBook(self,excelPathAndName):

        try:
            self.workbook=openpyxl.load_workbook(excelPathAndName)
        except Exception as e:
            raise e
        self.excelFile=excelPathAndName
        return self.workbook

    def getSheetNames(self):
        try:
            return self.workbook.sheetnames
        except Exception as e:
            raise e

    #根据sheet名获取该sheet对象
    def getSheetByName(self,sheetName):
        try:
            sheet=self.workbook.get_sheet_by_name(sheetName)
            return sheet
        except Exception as e:
            raise e


    #根据sheet的索引号获取该sheet对象
    def getSheetByIndex(self,sheetIndex):
        try:
            sheetname=self.workbook.get_sheet_names()[sheetIndex]
        except Exception as e:
            raise e
        sheet=self.workbook.get_sheet_by_name(sheetname)
        return sheet


    #获取sheet中有效数据区域的结束行号  最大行号
    def getRowsNumber(self,sheet):
        return sheet.max_row


    #获取sheet中有效数据区域的结束列号  最大列号
    def getColsNumber(self,sheet):
        return sheet.max_column


    #获取sheet中有效数据区域的开始行号 最小行号
    def getStartRowNumber(self,sheet):
        return sheet.min_row

    # 获取sheet中有效数据区域的开始列号 最小列号
    def getStartColNumber(self,sheet):
        return sheet.min_column



    '''
    获取sheet中某一行，返回的是这一行所有的数据内容组成的tuple，
    下标从1开始，sheet.rows[1]表示第一行
    '''
    def getRow(self,sheet,rowNo):
        try:
            return list(sheet.rows)[rowNo-1]
        except Exception as e:
            raise e



    '''
    获取sheet中某一列，返回的是这一列所有的数据内容组成的tuple，
    下标从1开始，sheet.rows[1]表示第一列
    '''
    def getColumn(self,sheet,colNo):
        try:
            return list(sheet.columns)[colNo-1]
        except Exception as e:
            raise e



    '''
    根据单元格所在的位置索引获取该单元格中的值，下标从1开始，
    sheet.cell(row=1,column=1).value,表示excel中第一行第一列的值
    '''
    def getCellOfValue(self,sheet,coordinate=None,rowNo=None,colsNo=None):
        if coordinate!=None:
            try:
                return sheet.cell(coordinate=coordinate).value
            except Exception as e:
                raise e
        elif coordinate is None and rowNo is not None and colsNo is not None:
            try:
                return sheet.cell(row=rowNo,column=colsNo).value
            except Exception as e:
                raise e
        else:
            raise Exception("Insufficient Coordinates od cell。。。")




    '''
    获取某个单元格的对象，可以根据单元格所在位置的数字索引，
    也可以直接根据Excel中单元格的编码及坐标
    如getCellObject(sheet,coordinate='A1')或
    getCellObject(sheet,rowNo=1,colsNo=2)
    '''
    def getCellOfObject(self,sheet,coordinate=None,rowNo=None,colsNo=None):
        if coordinate!=None:
            try:
                return sheet.cell(coordinate=coordinate)
            except Exception as e:
                raise e
        elif coordinate is None and rowNo is not None and colsNo is not None:
            try:
                return sheet.cell(row=rowNo,column=colsNo)
            except Exception as e:
                raise e
        else:
            raise Exception("Insufficient Coordinates od cell。。。")



    '''
    根据单元格在excel中的编码坐标或者数字索引坐标向单元格中写入数据，
    下标从1开始，参数style表示字体的颜色的名字，比如red，green
    '''
    def writeCell(self,sheet,content,coordinate=None,rowNo=None,colsNo=None,style=None):
        if coordinate is not None:
            try:
                sheet.cell(coordinate=coordinate).value=content
                if style is not None:
                    sheet.cell(coordinate=coordinate).font=Font(color=self.RGBDict[style])
                self.workbook.save(self.excelFile)
            except Exception as e:
                raise e
        elif coordinate==None and rowNo is not None and colsNo is not None:
            try:
                sheet.cell(row=rowNo,column=colsNo).value=content
                if style:
                    sheet.cell(row=rowNo,column=colsNo).font=Font(color=self.RGBDict[style])
                self.workbook.save(self.excelFile)
            except Exception as e:
                raise e
        else:
            raise Exception("Insufficient Coordinates od cell。。。")

# ParseExcel.writeCell(sheet="ces", content=getCellOfObject(), coordinate=F50)


    '''
    写入当前的时间，下标从1开始
    '''
    def writeCellCurrentTime(self,sheet,coordinate=None,rowNo=None,colsNo=None):
        now=int(time.time())
        timeArray=time.localtime(now)
        currentTime=time.strftime("%Y-%m-%d %H:%M:%S",timeArray)

        if coordinate is not None:
            try:
                sheet.cell(coordinate=coordinate).value=currentTime
                self.workbook.save(self.excelFile)
            except Exception as e:
                raise e
        elif coordinate==None and rowNo is not None and colsNo is not None:
            try:
                sheet.cell(row=rowNo,column=colsNo).value=currentTime
                self.workbook.save(self.excelFile)
            except Exception as e:
                raise e
        else:
            raise Exception("Insufficient Coordinates od cell。。。")




