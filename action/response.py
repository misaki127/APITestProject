#coding:utf-8
import requests
import re
from util.Log import *
from VarConfig import *
from con_Mysql import con_Mysql
from openpyxl.styles import Alignment,Font


def getResponse(url,method, **kwargs):
        # """封装request方法"""
        # # 获取请求参数
    params = kwargs.get("params")
    data = kwargs.get("data")
    json = kwargs.get("json")
    headers = kwargs.get('headers')
    cookies = kwargs.get('cookies')
    files = kwargs.get('files') #{name ,(filename,fileobj,'content_type', custom_headers) }
    auth = kwargs.get('auth')  #自定义身份验证
    timeout = kwargs.get('timeout') #超时
    allow_redirects = kwargs.get('allow_redirects')#boolen 是否运行重定向
    proxies = kwargs.get('proxies')#代理
    verify = kwargs.get('verify')#boolen 它控制我们是否验证服务器的TLS证书或字符串，在这种情况下，它必须是路径要使用的CA包。默认为“True”。
    stream = kwargs.get('stream')#如果``False``，则立即下载响应内容。
    cert = kwargs.get('cert') #如果是字符串，就是证书路径，如果是元组就是（证书，密钥）
    hooks = kwargs.get('hooks')#信号事件处理  传递一个 {hook_name: callback_function} 字典给 hooks 请求参数若执行你的回调函数期间发生错误，系统会给出一个警告。若回调函数返回一个值，默认以该值替换传进来的数据。若函数未返回任何东西， 也没有什么其他的影响

    try:
        if verify == None or verify == "":
            verify = False

        r = requests.request(method=method,url=url,params=params,data=data,json=json,headers=headers,cookies=cookies
                                     ,files=files,auth=auth,timeout=timeout,allow_redirects=allow_redirects,proxies=proxies
                                     ,verify=verify,stream=stream,cert=cert,hooks=hooks)
        return r
    except Exception as e:

        logging.info("请求错误: %s" % e)



def jsonGetInfo(data,findData,dataName):  #在返回的json数据里查找对应数据,返回对应的值，str格式，data:接口返回的json数据或str数据，endData:查找的value的KEY值，相同的名称可以只填一个,未找到返回一个None，dataName:找到数据后存入的变量名
    #                                                 #数据和变量名按顺序配对
    try:
        resultList = []  #查找结果数据
        if isinstance(data,str):
            pass
        else:
            data=str(data)
        data = data.strip()
        findData = findData.strip()
        dataName = dataName.strip()
        dataList = re.split('[^\w_-]',data)
        dataList = list(filter(None, dataList))
        findDataList = re.split(varibleSep,findData)
        dataNameList = re.split(varibleSep,dataName)
        for i in dataList:
            for a in findDataList:
                if a == i:
                    index = [c for c, x in enumerate(dataList) if x == i]

                    for b in index:
                        result = dataList[b + 1]
                        resultList.append(a)
                        resultList.append(result)

        resultDict = {}
        endDict = {}
        for f in range(0, len(resultList), 2):
            resultDict[resultList[f]] = resultList[f + 1]
        if len(resultDict) != len(findDataList):
            noList = []
            for k in findDataList:
                if k not in resultDict:
                    noList.append(k)
            logging.info(str(noList) + '值未取到')
        for p in resultDict:
            Pindex = [c for c, x in enumerate(findDataList) if x == p]
            endDict[dataNameList[Pindex[0]]] = resultDict[p]
        return endDict

    except IndexError:
        logging.info('查找' + str(a) + '时，查找到的值为空，请检查输入是否正确！')
    except Exception as e:
        logging.info('jsonGetInfo:' + str(e))



def jsonGetFirstInfo(data,findData,dataName):  #在json数据内查找指定ID值，并保存第一个找到的值，如未找到则提示输入查找值不正确
    try:
        resultList = []  #查找结果数据
        if isinstance(data,str):
            pass
        else:
            data=str(data)
        data = data.strip()
        findData = findData.strip()
        dataName = dataName.strip()
        dataList = re.split('[^\w_-]',data)
        dataList = list(filter(None, dataList))  #拆解json成列表
        findDataList = re.split(varibleSep,findData)  #查找值列表
        dataNameList = re.split(varibleSep,dataName)   #命名值列表
        for a in findDataList:
            for b in dataList:
                if a == b:
                    resultList.append(dataList[dataList.index(b)+1])
                    break

        resultDict = {}
        if  len(dataNameList) == len(resultList):

            for f in range(len(dataNameList)):
                resultDict[dataNameList[f]]=resultList[f]
            return resultDict
        else:
            logging.info('获取参数失败！请检查需要查找的数据与变量数是否一致!')
            raise ValueError

    except Exception as e:
        logging.info('jsonGetFirstInfo:' + str(e))



def splitCode(dataType,data):         #请求方式以逗号分隔，请求参数以$$$分隔 将请求方式与参数一一配对组合在一起，组成一个str
    try:
        if data =="" or data ==None or dataType =='' or dataType ==None:
            return None
        dataType = dataType.strip()
        data = data.strip()
        methodList = re.split(dataTypeSep,dataType)
        dataList = re.split(dataSep,data)
        if len(methodList) != len(dataList):

            logging.info('请求方式与请求参数不匹配，请核对后重试')
            result = None
        elif len(methodList) == 1 and methodList[0] == '':
            result = ''
        else:
            result = ''
            for a in range(len(methodList)):
                result +=methodList[a] + '=' + dataList[a] +','
        return result
    except Exception as e:

        logging.info('拼接代码：splitCode:' + str(e))


#适用于json格式的变量替换
def updateVaribleForDict(data,dict,separtor):      #从参数里确定变量位置和变量name，从变量字典内拿取数据填入
    try:
        for i in dict:
            data = data.replace(separtor+i+separtor,'"'+str(dict[i])+'"')  #变量用配置文件的符号包裹
            #data = data.replace(separtor + i + separtor, str(dict[i]))
        return data
    except Exception as e:

        logging.info('查找变量名填入变量值:' + str(e))


#适用于字符串的变量替换
def updateVaribleForStr(data,dict,separtor):      #从参数里确定变量位置和变量name，从变量字典内拿取数据填入
    try:
        for i in dict:
              #变量用配置文件的符号包裹
            data = data.replace(separtor + i + separtor, str(dict[i]))
        return data
    except Exception as e:

        logging.info('查找变量名填入变量值:' + str(e))


def sqlGetVarible(sql, varible):  #查询sql获取结果，并将结果与变量名一一对应，返回一个字典，如数量不对应则返回一个None
    try:
        sqlList = re.split(sqlSeq,sql)
        mysql = con_Mysql()
        p = []
        for i in sqlList:
            mysql.sql = i
            logging.info('sql is {0}'.format(i))

            r = mysql.select_sql()
            for a in r:
                for b in a:
                    if a[b] in p:
                        pass
                    else:
                        p.append(a[b])
            logging.info('sql results is {0}'.format(str(p)))

        mysql.end_con()
        varibleList = re.split(varibleSep,varible)
        varibleDict = {}
        if len(varibleList) == len(p):
            for l in range(len(p)):
                varibleDict[varibleList[l]] = p[l]

        else:
            # print('SQL返回的数据与设置的变量名数量不对应，请检查是否输入正确！')
            logging.info('SQL返回的数据与设置的变量名数量不对应，请检查是否输入正确！')
            varibleDict = None
        return varibleDict
    except Exception as e:
        logging.info('sql获取参数错误:'+str(e))



import time
def getImportInfo(r):  #传入请求响应，自动写入excel
    try:
        file =importFilePath+str(int(time.time()))+'.xlsx'
        f = open(file,'wb')
        f.write(r.content)
        f.close()
        return file
    except Exception as e:
        logging.info("获取导出数据错误 error is "+str(e))

import openpyxl


def getExcelObj(filePath):
    try:
        return openpyxl.load_workbook(filePath)
    except Exception as e:
        logging.info(str(e))

def createReportSheet(filePath,data):
    #初始化接口自动化报告
    try:
        excelObj=getExcelObj(filePath)
        sheetList = excelObj.sheetnames
        if testReportSheetName in sheetList:
            del excelObj[testReportSheetName]
        excelObj.create_sheet(title=testReportSheetName)
        reportSheet = excelObj[testReportSheetName]
        reportSheet['A1'].alignment = Alignment(horizontal='center', vertical='center')
        list1 = [{'sheetName':testReportSheetName,'address':'A1','value':'接口自动化脚本' + '测试报告'},
                 {'sheetName':testReportSheetName,'address':'A2','value':'开始时间'},
                 {'sheetName':testReportSheetName,'address':'C2','value':'耗时'},
                 {'sheetName':testReportSheetName,'address':'E2','value':'结果'},
                 {'sheetName': testReportSheetName, 'address': 'A3', 'value': '用例组名称/用例名称'},
                 {'sheetName': testReportSheetName, 'address': 'B3', 'value': '总数'},
                 {'sheetName': testReportSheetName, 'address': 'C3', 'value': '通过'},
                 {'sheetName': testReportSheetName, 'address': 'D3', 'value': '失败'},
                 {'sheetName': testReportSheetName, 'address': 'E3', 'value': '错误'},
                 {'sheetName': testReportSheetName, 'address': 'F3', 'value': '结果信息/错误信息'},
                 {'sheetName': testReportSheetName, 'address': 'A1:F1', 'value': 'merge_cells'}
                 ]
        data = data + list1
        for i in data:
            if i['value'] == 'merge_cells':
                sheet = excelObj[i['sheetName']]
                sheet.merge_cells(i['address'])
            else:
                sheet = excelObj[i['sheetName']]
                sheet[i['address']] = i['value']
            if i['value'] == 'pass' or i['value']== 200 or i['value']=='200':
                sheet[i['address']].alignment = Alignment(horizontal='center',vertical='center')
                sheet[i['address']].font = Font(color='00FF00')
            elif i['value'] == 'faild' or i['value']== 500 or i['value']=='500':
                sheet[i['address']].alignment = Alignment(horizontal='center', vertical='center')
                sheet[i['address']].font = Font(color='DC143C')
        excelObj.save(dataFilePath)
    except Exception as e:
        logging.info(str(e))



#搜索功能  查找字段里，输入key:value,key1,key2 则可以自动寻找key:value对应的数据，然后从数据里查找需要的变量，需要返回的数据为{“results":{"list":{[]}}}格式，可输入多个字段搜索
def jsonSearch(jsonData, findData, dataName):
    try:
        findData = findData.strip()
        dataName = dataName.strip()
        findDataList = re.split(varibleSep, findData)  # 查找值列表
        dataNameList = re.split(varibleSep, dataName)  # 命名值列表
        searchList = []
        remList = []
        for i in findDataList:
            if ":" in i:
                searchList = searchList + re.split(":", i)
                remList.append(i)
        findDataList = [x for x in findDataList if x not in remList]
        endDict = {}
        if isinstance(jsonData, dict):
            jsonList = jsonData['results']['list']
            for data in jsonList:
                for index in range(0,len(searchList),2):
                    keyList = list(data.keys())
                    if searchList[index] not in keyList:
                        endDict = {}
                        break
                    if str(data[searchList[index]]) != str(searchList[index+1]):
                        endDict = {}
                        break
                    else:
                        endDict = data

                if endDict != {}:
                    break
            if isinstance(endDict, str):
                pass
            else:
                endDict = str(endDict)
            dataList = re.split('[^\w_-]', endDict)
            #dataList = list(filter(None, dataList))  # 拆解json成列表
            resultList = []
            for a in findDataList:
                for b in dataList:
                    if a == b:
                        resultList.append(dataList[dataList.index(b) + 1])
                        break

            resultDict = {}
            if len(dataNameList) == len(resultList):

                for f in range(len(dataNameList)):
                    resultDict[dataNameList[f]] = resultList[f]
                return resultDict
            else:
                logging.info('获取参数失败！请检查数据是否合法!')
                raise ValueError

        else:
            raise SyntaxError

    except Exception as e:
        logging.info("搜索功能失败："+e)


#输入预期结果时，对结果进行检测
def passTesting(jsonData,data):
    try:
        dataList = re.split(expectResultSep,str(data))
        for i in dataList:
            if i in str(jsonData):
                pass
            else:
                return False
        return True
    except Exception as e:
        logging.debug('jsonData:'+str(jsonData) + ',data = ' + str(data))
        logging.info("通过检测失败！")